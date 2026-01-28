import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv

# Читаем CSV
data = []
with open('SEMANTIKA_SECONDHAND_VINTAGE.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter='\t')
    for row in reader:
        data.append(row)

# Создаём Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Семантика'

# Добавляем данные
for row_idx, row in enumerate(data, 1):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Форматирование заголовка
        if row_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF', size=12)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border

# Ширина колонок
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 45

# Высота заголовка
ws.row_dimensions[1].height = 25

# Сохраняем
wb.save('SEMANTIKA_SECONDHAND_VINTAGE.xlsx')
print('Excel файл создан: SEMANTIKA_SECONDHAND_VINTAGE.xlsx')
